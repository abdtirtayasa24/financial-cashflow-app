"""Behavior tests for approved-only reporting APIs."""

from __future__ import annotations

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from tests.conftest import auth_header, make_token, seed_user
from tests.fakes import FakeClient

DEPT_PARENT = "dept-parent"
DEPT_CHILD = "dept-child"
DEPT_OTHER = "dept-other"
CAT_PARENT = "cat-parent"
CAT_CHILD = "cat-child"
CASH_MAIN = "cash-main"
CASH_SECOND = "cash-second"


def seed_report_row(
    db: FakeClient,
    *,
    row_id: str,
    transaction_date: str,
    direction: str,
    base_amount: float,
    department_id: str = DEPT_PARENT,
    department_name: str = "Parent Dept",
    category_id: str = CAT_PARENT,
    category_name: str = "Parent Category",
    cash_account_id: str = CASH_MAIN,
    cash_account_name: str = "Main Bank",
    transaction_no: str | None = None,
) -> None:
    db.seed(
        "approved_cashflow_report_base",
        [
            {
                "id": row_id,
                "transaction_no": transaction_no or row_id,
                "transaction_date": transaction_date,
                "direction": direction,
                "amount": base_amount,
                "base_amount": base_amount,
                "currency": "IDR",
                "cash_account_id": cash_account_id,
                "cash_account_name": cash_account_name,
                "department_id": department_id,
                "department_name": department_name,
                "department_code": department_id.upper(),
                "category_id": category_id,
                "category_name": category_name,
                "payment_method_id": "pm-1",
                "payment_method_name": "Transfer",
                "counterparty_name": None,
                "reference_no": None,
                "description": None,
                "status": "APPROVED",
                "reviewed_at": "2026-07-01T00:00:00+00:00",
            }
        ],
    )


def seed_cash_account(
    db: FakeClient,
    *,
    account_id: str,
    name: str,
    opening_balance: float,
    opening_balance_date: str,
) -> None:
    db.seed(
        "cash_accounts",
        [
            {
                "id": account_id,
                "name": name,
                "account_type": "BANK",
                "opening_balance": opening_balance,
                "opening_balance_date": opening_balance_date,
                "currency": "IDR",
                "is_active": True,
                "created_at": "2026-01-01T00:00:00+00:00",
            }
        ],
    )


def seed_transaction(
    db: FakeClient,
    *,
    transaction_id: str,
    status: str,
    department_id: str = DEPT_PARENT,
    transaction_date: str = "2026-07-01",
) -> None:
    db.seed(
        "cashflow_transactions",
        [
            {
                "id": transaction_id,
                "transaction_no": transaction_id,
                "transaction_date": transaction_date,
                "direction": "INFLOW",
                "amount": 100,
                "currency": "IDR",
                "exchange_rate": 1,
                "base_amount": 100,
                "cash_account_id": CASH_MAIN,
                "department_id": department_id,
                "category_id": CAT_PARENT,
                "payment_method_id": "pm-1",
                "counterparty_name": None,
                "reference_no": None,
                "description": None,
                "status": status,
                "created_by": "u1",
                "submitted_at": None,
                "reviewed_by": None,
                "reviewed_at": None,
                "rejection_reason": None,
                "void_reason": None,
                "created_at": "2026-07-01T00:00:00+00:00",
                "updated_at": "2026-07-01T00:00:00+00:00",
            }
        ],
    )


def report_token(fake_db: FakeClient, role: str = "FINANCE_ADMIN") -> dict[str, str]:
    user_id = seed_user(fake_db, role, email=f"{role.lower()}@example.com")
    return auth_header(make_token(user_id))


def test_summary_includes_only_approved_view_rows_and_filters(
    client: TestClient, fake_db: FakeClient
) -> None:
    token = report_token(fake_db)
    seed_report_row(
        fake_db, row_id="in-1", transaction_date="2026-07-10",
        direction="INFLOW", base_amount=1_000, department_id=DEPT_PARENT,
    )
    seed_report_row(
        fake_db, row_id="out-1", transaction_date="2026-07-11",
        direction="OUTFLOW", base_amount=250, department_id=DEPT_PARENT,
    )
    seed_report_row(
        fake_db, row_id="old", transaction_date="2026-06-30",
        direction="INFLOW", base_amount=9_999, department_id=DEPT_PARENT,
    )
    seed_report_row(
        fake_db, row_id="other-dept", transaction_date="2026-07-10",
        direction="INFLOW", base_amount=777, department_id=DEPT_OTHER,
    )
    # Non-approved statuses live in the transaction table and must not affect reports.
    for status in ["DRAFT", "SUBMITTED", "REJECTED", "VOIDED"]:
        seed_transaction(fake_db, transaction_id=f"tx-{status}", status=status)

    resp = client.get(
        "/api/reports/summary?date_from=2026-07-01&date_to=2026-07-31"
        f"&department_id={DEPT_PARENT}",
        headers=token,
    )

    assert resp.status_code == 200, resp.text
    assert resp.json() == {
        "totalInflow": 1000.0,
        "totalOutflow": 250.0,
        "netCashflow": 750.0,
        "currency": "IDR",
    }


def test_monthly_trend_groups_approved_rows_oldest_first(
    client: TestClient, fake_db: FakeClient
) -> None:
    token = report_token(fake_db)
    seed_report_row(fake_db, row_id="aug-out", transaction_date="2026-08-03",
                    direction="OUTFLOW", base_amount=50)
    seed_report_row(fake_db, row_id="jul-in", transaction_date="2026-07-03",
                    direction="INFLOW", base_amount=300)
    seed_report_row(fake_db, row_id="jul-out", transaction_date="2026-07-04",
                    direction="OUTFLOW", base_amount=125)

    resp = client.get("/api/reports/monthly-trend", headers=token)

    assert resp.status_code == 200
    assert resp.json() == [
        {"month": "2026-07", "inflow": 300.0, "outflow": 125.0, "net": 175.0},
        {"month": "2026-08", "inflow": 0.0, "outflow": 50.0, "net": -50.0},
    ]


def test_category_and_department_breakdowns_are_exact_match_not_rollup(
    client: TestClient, fake_db: FakeClient
) -> None:
    token = report_token(fake_db)
    seed_report_row(
        fake_db, row_id="parent", transaction_date="2026-07-01",
        direction="OUTFLOW", base_amount=100, department_id=DEPT_PARENT,
        department_name="Parent Dept", category_id=CAT_PARENT,
        category_name="Parent Category",
    )
    seed_report_row(
        fake_db, row_id="child", transaction_date="2026-07-01",
        direction="OUTFLOW", base_amount=999, department_id=DEPT_CHILD,
        department_name="Child Dept", category_id=CAT_CHILD,
        category_name="Child Category",
    )

    category = client.get(
        f"/api/reports/by-category?category_id={CAT_PARENT}", headers=token
    )
    department = client.get(
        f"/api/reports/by-department?department_id={DEPT_PARENT}", headers=token
    )

    assert category.status_code == 200
    assert category.json() == [
        {
            "category_id": CAT_PARENT,
            "category_name": "Parent Category",
            "direction": "OUTFLOW",
            "amount": 100.0,
        }
    ]
    assert department.status_code == 200
    assert department.json() == [
        {
            "department_id": DEPT_PARENT,
            "department_name": "Parent Dept",
            "inflow": 0.0,
            "outflow": 100.0,
            "net": -100.0,
        }
    ]


def test_cash_account_balance_ignores_date_range_and_opening_date_rules_apply(
    client: TestClient, fake_db: FakeClient
) -> None:
    token = report_token(fake_db)
    seed_cash_account(
        fake_db, account_id=CASH_MAIN, name="Main Bank",
        opening_balance=1_000, opening_balance_date="2026-07-01",
    )
    seed_report_row(
        fake_db, row_id="before-opening", transaction_date="2026-06-30",
        direction="INFLOW", base_amount=5_000, cash_account_id=CASH_MAIN,
        cash_account_name="Main Bank",
    )
    seed_report_row(
        fake_db, row_id="after-in", transaction_date="2026-07-15",
        direction="INFLOW", base_amount=300, cash_account_id=CASH_MAIN,
        cash_account_name="Main Bank",
    )
    seed_report_row(
        fake_db, row_id="after-out", transaction_date="2026-08-15",
        direction="OUTFLOW", base_amount=125, cash_account_id=CASH_MAIN,
        cash_account_name="Main Bank",
    )

    resp = client.get(
        "/api/reports/cash-account-balances?date_from=2026-09-01&date_to=2026-09-30",
        headers=token,
    )

    assert resp.status_code == 200, resp.text
    assert resp.json() == [
        {
            "cash_account_id": CASH_MAIN,
            "cash_account_name": "Main Bank",
            "currency": "IDR",
            "opening_balance": 1000.0,
            "inflow": 300.0,
            "outflow": 125.0,
            "current_balance": 1175.0,
        }
    ]


def test_pending_approval_count_counts_submitted_only_with_filters(
    client: TestClient, fake_db: FakeClient
) -> None:
    token = report_token(fake_db, "MANAGEMENT")
    seed_transaction(
        fake_db, transaction_id="s1", status="SUBMITTED",
        department_id=DEPT_PARENT, transaction_date="2026-07-15",
    )
    seed_transaction(
        fake_db, transaction_id="s-old", status="SUBMITTED",
        department_id=DEPT_PARENT, transaction_date="2026-06-30",
    )
    seed_transaction(
        fake_db, transaction_id="s2", status="SUBMITTED",
        department_id=DEPT_OTHER, transaction_date="2026-07-15",
    )
    seed_transaction(
        fake_db, transaction_id="a1", status="APPROVED",
        department_id=DEPT_PARENT, transaction_date="2026-07-15",
    )

    resp = client.get(
        f"/api/reports/pending-approvals?department_id={DEPT_PARENT}"
        "&date_from=2026-07-01&date_to=2026-07-31",
        headers=token,
    )

    assert resp.status_code == 200
    assert resp.json() == {"count": 1}


def test_employee_and_system_admin_cannot_access_reports(
    client: TestClient, fake_db: FakeClient
) -> None:
    employee = report_token(fake_db, "EMPLOYEE")
    system_admin = report_token(fake_db, "SYSTEM_ADMIN")

    assert client.get("/api/reports/summary", headers=employee).status_code == 403
    assert client.get("/api/reports/summary", headers=system_admin).status_code == 403


def test_export_summary_xlsx_and_pdf_use_report_totals(
    client: TestClient, fake_db: FakeClient
) -> None:
    token = report_token(fake_db)
    seed_report_row(fake_db, row_id="in-1", transaction_date="2026-07-01",
                    direction="INFLOW", base_amount=1000)
    seed_report_row(fake_db, row_id="out-1", transaction_date="2026-07-02",
                    direction="OUTFLOW", base_amount=250)

    summary = client.get("/api/reports/summary", headers=token)
    xlsx = client.post(
        "/api/reports/export",
        headers=token,
        json={"format": "xlsx", "report_type": "summary"},
    )
    pdf = client.post(
        "/api/reports/export",
        headers=token,
        json={"format": "pdf", "report_type": "summary"},
    )

    assert summary.status_code == 200
    assert xlsx.status_code == 200, xlsx.text
    assert xlsx.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert xlsx.content.startswith(b"PK")
    workbook = load_workbook(BytesIO(xlsx.content), data_only=True)
    sheet = workbook.active
    exported = {sheet.cell(row=i, column=1).value: sheet.cell(row=i, column=2).value
                for i in range(2, 5)}
    body = summary.json()
    assert exported["Total inflow"] == body["totalInflow"]
    assert exported["Total outflow"] == body["totalOutflow"]
    assert exported["Net cashflow"] == body["netCashflow"]
    assert pdf.status_code == 200, pdf.text
    assert pdf.headers["content-type"].startswith("application/pdf")
    assert pdf.content.startswith(b"%PDF")
