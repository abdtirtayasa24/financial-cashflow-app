from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Any
from zipfile import BadZipFile

from openpyxl import load_workbook  # type: ignore[import-untyped]
from openpyxl.utils.exceptions import InvalidFileException  # type: ignore[import-untyped]
from supabase import Client

from app.core.errors import AppError
from app.core.models import CurrentUser, Role
from app.modules.imports.repository import TransactionImportRepository
from app.modules.imports.schemas import (
    ImportedTransactionOut,
    ImportRowError,
    ImportTransactionsOut,
)
from app.modules.transactions.service import editable_snapshot

EXPECTED_COLUMNS = [
    "transaction_date",
    "direction",
    "amount",
    "category_name",
    "department_code",
    "cash_account_name",
    "payment_method_name",
    "counterparty_name",
    "reference_no",
    "description",
]
REQUIRED_COLUMNS = EXPECTED_COLUMNS[:7]
MAX_IMPORT_ROWS = 500
MAX_IMPORT_FILE_BYTES = 10 * 1024 * 1024


@dataclass
class ParsedRow:
    row_number: int
    values: dict[str, Any]


class TransactionImportService:
    def __init__(self, db: Client) -> None:
        self.repo = TransactionImportRepository(db)

    def import_file(
        self,
        *,
        filename: str,
        content: bytes,
        user: CurrentUser,
    ) -> ImportTransactionsOut:
        if user.role not in {Role.FINANCE_ADMIN, Role.MANAGEMENT}:
            raise AppError("forbidden", 403)
        if len(content) > MAX_IMPORT_FILE_BYTES:
            raise AppError("Import file exceeds 10 MB size limit", 422)
        rows = self._parse(filename, content)
        imported: list[ImportedTransactionOut] = []
        errors: list[ImportRowError] = []
        for parsed in rows:
            try:
                tx = self._import_row(parsed.values, user)
                imported.append(
                    ImportedTransactionOut(
                        id=tx["id"], transaction_no=tx["transaction_no"]
                    )
                )
            except AppError as exc:
                errors.append(
                    ImportRowError(row_number=parsed.row_number, message=exc.message)
                )
        return ImportTransactionsOut(
            total_rows=len(rows),
            imported_count=len(imported),
            failed_count=len(errors),
            imported=imported,
            errors=errors,
        )

    def _parse(self, filename: str, content: bytes) -> list[ParsedRow]:
        lower = filename.lower()
        if lower.endswith(".csv"):
            return self._parse_csv(content)
        if lower.endswith(".xlsx"):
            return self._parse_xlsx(content)
        raise AppError("Only CSV and XLSX files are supported", 422)

    def _parse_csv(self, content: bytes) -> list[ParsedRow]:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError as exc:
            raise AppError("Invalid CSV file encoding", 422) from exc
        reader = csv.DictReader(StringIO(text))
        self._require_columns(list(reader.fieldnames or []))
        out: list[ParsedRow] = []
        for i, row in enumerate(reader, start=2):
            if any(_clean(v) for v in row.values()):
                out.append(
                    ParsedRow(
                        row_number=i,
                        values={k: _clean(v) for k, v in row.items()},
                    )
                )
                self._enforce_row_limit(len(out))
        return out

    def _parse_xlsx(self, content: bytes) -> list[ParsedRow]:
        try:
            workbook = load_workbook(BytesIO(content), data_only=True, read_only=True)
        except (BadZipFile, InvalidFileException, OSError, ValueError) as exc:
            raise AppError("Invalid XLSX file", 422) from exc
        try:
            sheet = workbook.active
            rows_iter = sheet.iter_rows(values_only=True)
            header = [str(value or "").strip() for value in next(rows_iter, ())]
            self._require_columns(header)
            out: list[ParsedRow] = []
            for idx, values in enumerate(rows_iter, start=2):
                row = {
                    header[i]: _clean(values[i]) if i < len(values) else ""
                    for i in range(len(header))
                }
                if any(row.values()):
                    out.append(ParsedRow(row_number=idx, values=row))
                    self._enforce_row_limit(len(out))
            return out
        finally:
            workbook.close()

    def _require_columns(self, columns: list[str]) -> None:
        missing = [column for column in REQUIRED_COLUMNS if column not in columns]
        if missing:
            raise AppError(f"Missing required columns: {', '.join(missing)}", 422)

    def _enforce_row_limit(self, count: int) -> None:
        if count > MAX_IMPORT_ROWS:
            raise AppError("Import file exceeds 500 row limit", 422)

    def _import_row(self, row: dict[str, Any], user: CurrentUser) -> dict[str, Any]:
        values = {column: row.get(column, "") for column in EXPECTED_COLUMNS}
        for column in REQUIRED_COLUMNS:
            if values[column] in (None, ""):
                raise AppError(f"{column} is required", 422)
        transaction_date = _date_str(values["transaction_date"])
        direction = str(values["direction"]).strip().upper()
        if direction not in {"INFLOW", "OUTFLOW"}:
            raise AppError("direction must be INFLOW or OUTFLOW", 422)
        try:
            amount = float(values["amount"])
        except (TypeError, ValueError) as exc:
            raise AppError("amount must be numeric", 422) from exc
        if amount <= 0:
            raise AppError("amount must be positive", 422)
        department = self.repo.department_by_code(str(values["department_code"]).strip())
        if not department:
            raise AppError(f"Unknown department code '{values['department_code']}'", 422)
        categories = self.repo.categories_by_name(str(values["category_name"]).strip())
        if not categories:
            raise AppError(f"Unknown category '{values['category_name']}'", 422)
        category = next(
            (c for c in categories if c.get("direction") in {direction, "BOTH"}),
            None,
        )
        if not category:
            raise AppError("Category direction is incompatible", 422)
        cash_account_name = str(values["cash_account_name"]).strip()
        cash_accounts = self.repo.cash_accounts_by_name(cash_account_name)
        if not cash_accounts:
            raise AppError(f"Unknown cash account '{values['cash_account_name']}'", 422)
        if len(cash_accounts) > 1:
            raise AppError(
                f"Cash account name '{values['cash_account_name']}' is ambiguous", 422
            )
        cash_account = cash_accounts[0]
        payment_method = self.repo.payment_method_by_name(
            str(values["payment_method_name"]).strip()
        )
        if not payment_method:
            raise AppError(
                f"Unknown payment method '{values['payment_method_name']}'", 422
            )
        payload = {
            "transaction_no": self.repo.transactions.next_transaction_no(
                direction, transaction_date
            ),
            "transaction_date": transaction_date,
            "direction": direction,
            "amount": amount,
            "currency": "IDR",
            "exchange_rate": 1.0,
            "base_amount": amount,
            "cash_account_id": cash_account["id"],
            "department_id": department["id"],
            "category_id": category["id"],
            "payment_method_id": payment_method["id"],
            "counterparty_name": _optional(values["counterparty_name"]),
            "reference_no": _optional(values["reference_no"]),
            "description": _optional(values["description"]),
            "status": "DRAFT",
            "created_by": user.id,
        }
        tx = self.repo.transactions.insert(payload)
        self.repo.transactions.audit(
            tx["id"], user.id, "CREATE", new_value=editable_snapshot(tx)
        )
        return tx


def _clean(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def _optional(value: Any) -> str | None:
    value = _clean(value)
    return None if value == "" else str(value)


def _date_str(value: Any) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    try:
        return date.fromisoformat(str(value)).isoformat()
    except ValueError as exc:
        raise AppError("transaction_date must be a valid date", 422) from exc
